# -*- coding: utf-8 -*-
'''
@Time    : 2023/3/24 0024 上午 10:33
@Author  : fengaijun
@File    : ExcelAction.py

'''

import openpyxl
#创建Excel
def createExcel():
    # 创建工作簿
    wk = openpyxl.Workbook()
    # 获取当前工作表
    sheet = wk.active
    #写入数据至单元格
    sheet.cell(1,1,"username")
    sheet.cell(1,2,"class")
    sheet.cell(1,3,"score")
    wk.save("userinfo.xlsx")

#读取Excel
def readExcel(filepath):
    #加载工作簿
    wk = openpyxl.load_workbook(filepath)
    #获取工作表
    Sheet1 = wk["Sheet"]
    # Sheet1 = wk.get_sheet_names("Sheet")   #官方不建议使用
    #获取单元格数据
    #坐标
    location = Sheet1.cell(1,1)
    #单元格-值
    val = Sheet1.cell(1,1).value
    print(location,val)
    #获取表行数
    rows = Sheet1.max_row
    #获取表列数
    cols = Sheet1.max_column
    vallist = []
    for row in range(2,rows+1):
        for col in range(1,cols+1):
            value = Sheet1.cell(row,col).value
            print(value)
            vallist.append(value)
    print(vallist)

#编辑Excel
def EditExcel(filepath):
    #加载工作簿
    wk = openpyxl.load_workbook(filepath)
    #创建新的工作表
    Sheet2 = wk.create_sheet("我爱一条柴",2)
    Sheet2.cell(1,1,"韦小宝")
    Sheet2.cell(1,2,"多隆")
    Sheet2.cell(1,3,"陈近南")
    wk.save(filepath)

def deleteExcel(filepath,sheetname):
    #加载工作簿
    wk = openpyxl.load_workbook(filepath)
    #删除指定工作表
    wk.remove(wk[sheetname])
    #del wk["mysheet"]
    wk.save(filepath)

'''
一次性插入多个数据：
'''
def addManyData(filepath,sheetname):
    wk = openpyxl.load_workbook(filepath)
    if sheetname not in wk:
       #创建新的工作表
       newsheet= wk.create_sheet(sheetname)
       # 定义要插入的数据
       data = [
           [1, 'apple', 2],
           [2, 'orange', 3],
           [3, 'banana', 4]
       ]
       # 遍历数据并将它们插入到行中
       for row in data:
           newsheet.append(row)

       # 保存工作表
       wk.save(filepath)
    else:
        #获取工作表
        SheetName= wk[sheetname]
        # 定义要插入的数据
        data = [
            [1, 'apple', 2],
            [2, 'orange', 3],
            [3, 'banana', 4]
        ]
        # 遍历数据并将它们插入到行中
        for row in data:
            SheetName.append(row)

        # 保存工作表
        wk.save(filepath)
if __name__ == '__main__':
    # createExcel()
    # readExcel("userinfo.xlsx")
    # EditExcel("userinfo.xlsx")
    addManyData("userinfo.xlsx","伸腿瞪眼丸2")